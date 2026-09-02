from django.views.generic import TemplateView, CreateView, ListView, View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.contrib.auth.hashers import make_password, check_password
from django.utils import timezone
from django.core.validators import ValidationError
from django.urls import reverse_lazy
from django.shortcuts import redirect, render
from .models import Student, VideoLink, Signature
from .utils import preview_signature_on_template, generate_certificate_for_student
from .forms import StudentForm, ExcelUploadForm, CheckForm, SetPasswordForm, LoginPasswordForm, CertificateForm, PaymentForm
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse, Http404
from django.contrib.admin.views.decorators import staff_member_required
from datetime import datetime
from django.db import IntegrityError
 
class HomeView(TemplateView):
    template_name = 'home.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        student_id = self.request.session.get('auth_student_id')
        if student_id:
            try:
                student = Student.objects.get(pk=student_id)
                context['logged_in'] = True
                context['student'] = student
            except Student.DoesNotExist:
                # اگر دانشجو وجود نداشت، سشن رو پاک کن
                self.request.session.pop('auth_student_id', None)
                context['logged_in'] = False
                context['student'] = None
        else:
            context['logged_in'] = False
            context['student'] = None
            
        return context
 
class RegisterView(CreateView):
    model = Student
    form_class = StudentForm
    template_name = 'register.html'
    success_url = reverse_lazy('core:success')
    
    def form_valid(self, form):
        response = super().form_valid(form)
 
        # اطلاعات لازم برای صفحه‌ی «ثبت‌نام موفق»
        self.request.session['student_name'] = form.instance.name
        self.request.session['student_age'] = form.instance.age
        self.request.session['student_phone'] = form.instance.phone_number
        self.request.session['student_email'] = form.instance.email or ''
        self.request.session['student_reshte'] = form.instance.reshte
        self.request.session['student_school'] = form.instance.school
        self.request.session['student_city'] = form.instance.city
        self.request.session['student_moaref'] = form.instance.moaref or ''
 
        # چون دانشجو همین الان رمز عبور تعیین کرده، مستقیماً واردش می‌کنیم
        self.request.session['auth_student_id'] = form.instance.pk
 
        messages.success(self.request, f'✅ دانشجو  {form.instance.name} با موفقیت ثبت شد!')
        return response
    
    def form_invalid(self, form):
        if hasattr(form, 'existing_student') and form.existing_student:
            existing_student = form.existing_student
            # کاربر را به پنل کاربری هدایت کن
            self.request.session['auth_student_id'] = existing_student.pk
            self.request.session['student_name'] = existing_student.name
            self.request.session['student_age'] = existing_student.age
            self.request.session['student_phone'] = existing_student.phone_number
            self.request.session['student_email'] = existing_student.email or ''
            self.request.session['student_reshte'] = existing_student.reshte
            self.request.session['student_school'] = existing_student.school
            self.request.session['student_city'] = existing_student.city
            self.request.session['student_moaref'] = existing_student.moaref or ''
            
            messages.warning(self.request, f'⚠️ این شماره موبایل قبلاً برای دانشجو {existing_student.name} ثبت شده است. شما به پنل کاربری هدایت شدید.')
            return redirect('core:check_view')
        
        messages.error(self.request, 'خطا در ثبت‌نام! لطفاً اطلاعات را بررسی کنید.')
        return super().form_invalid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'ثبت‌نام دانش‌آموز'
        return context
 
 
class StudentsView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Student
    template_name = 'students.html'
    context_object_name = 'students'
    ordering = ['-created_at']
    
    def handle_no_permission(self):
        messages.error(self.request, 'شما دسترسی به این صفحه ندارید!')
        return redirect('core:home')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_count'] = Student.objects.count()
        context['certificate_count'] = Student.objects.filter(national_code__isnull=False).count()
        context['title'] = 'لیست دانش‌آموزان'
        return context
    
    def test_func(self):
        return self.request.user.is_staff
    
    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset
 
# ==========================================================================
# پنل کاربری با ورود واقعی (شماره/ایمیل + رمز عبور)
# ==========================================================================
 
class CheckView(View):
    """
    مرحله‌ی اول ورود: گرفتن شماره موبایل یا ایمیل.
    - اگر کاربر از قبل لاگین کرده باشد (auth_student_id در سشن)، مستقیم پنل نشان داده می‌شود.
    - اگر شماره/ایمیل معتبر باشد ولی کاربر هنوز رمز عبور نداشته باشد (کاربران قدیمی)،
      به صفحه‌ی «تعیین رمز عبور» هدایت می‌شود.
    - اگر رمز عبور داشته باشد، به صفحه‌ی «ورود با رمز عبور» هدایت می‌شود.
    """
    template_name = 'check.html'
 
    def get(self, request):
        # امکان خروج از مرحله‌ی رمز عبور و بازگشت به فرم اولیه با ?reset=1
        if request.GET.get('reset'):
            request.session.pop('pending_student_id', None)
 
        student_id = request.session.get('auth_student_id')
        if student_id:
            student = Student.objects.filter(pk=student_id).first()
            if student:
                return render(request, self.template_name, {
                    'found': True,
                    'student': student,
                    'logged_in': True,
                })
            request.session.pop('auth_student_id', None)
 
        form = CheckForm()
        return render(request, self.template_name, {'form': form})
 
    def post(self, request):
        form = CheckForm(request.POST)
        context = {'form': form}
 
        if form.is_valid():
            identifier = form.cleaned_data['identifier'].strip()
 
            try:
                if identifier.isdigit() and len(identifier) == 11:
                    student = Student.objects.get(phone_number=identifier)
                else:
                    student = Student.objects.get(email=identifier)
 
                # شناسه‌ی دانشجو را موقتاً در سشن نگه می‌داریم تا مرحله‌ی رمز عبور طی شود
                request.session['pending_student_id'] = student.pk
 
                if student.password:
                    return redirect('core:login_password')
                return redirect('core:set_password')
 
            except Student.DoesNotExist:
                context['found'] = False
                context['error'] = 'ایمیل یا شماره تماس پیدا نشد، با پشتیبانی ارتباط برقرار کنید'
 
        return render(request, self.template_name, context)

class PendingStudentMixin:
    """کمک‌کننده برای صفحات تعیین/ورود رمز عبور که به pending_student_id نیاز دارند."""
 
    def get_pending_student(self, request):
        student_id = request.session.get('pending_student_id')
        if not student_id:
            return None
        return Student.objects.filter(pk=student_id).first()
 
class SetPasswordView(PendingStudentMixin, View):
    """تعیین رمز عبور برای اولین بار (کاربرانی که قبل از این قابلیت ثبت‌نام کرده‌اند)."""
    template_name = 'check_password.html'
 
    def get(self, request):
        student = self.get_pending_student(request)
        if not student:
            messages.warning(request, 'ابتدا شماره موبایل یا ایمیل خود را در پنل کاربری وارد کنید.')
            return redirect('core:check_view')
        if student.password:
            return redirect('core:login_password')
 
        form = SetPasswordForm()
        return render(request, self.template_name, {'form': form, 'mode': 'set', 'student': student})
 
    def post(self, request):
        student = self.get_pending_student(request)
        if not student:
            messages.warning(request, 'ابتدا شماره موبایل یا ایمیل خود را در پنل کاربری وارد کنید.')
            return redirect('core:check_view')
 
        form = SetPasswordForm(request.POST)
        if form.is_valid():
            student.password = make_password(form.cleaned_data['password1'])
            student.save(update_fields=['password'])
 
            request.session.pop('pending_student_id', None)
            request.session['auth_student_id'] = student.pk
 
            messages.success(request, '✅ رمز عبور شما با موفقیت تنظیم شد و وارد پنل شدید.')
            return redirect('core:check_view')
 
        return render(request, self.template_name, {'form': form, 'mode': 'set', 'student': student})

class LoginPasswordView(PendingStudentMixin, View):
    """ورود با رمز عبور برای کاربرانی که قبلاً رمز تعیین کرده‌اند."""
    template_name = 'check_password.html'
 
    def get(self, request):
        student = self.get_pending_student(request)
        if not student:
            messages.warning(request, 'ابتدا شماره موبایل یا ایمیل خود را در پنل کاربری وارد کنید.')
            return redirect('core:check_view')
        if not student.password:
            return redirect('core:set_password')
 
        form = LoginPasswordForm()
        return render(request, self.template_name, {'form': form, 'mode': 'login', 'student': student})
 
    def post(self, request):
        student = self.get_pending_student(request)
        if not student:
            messages.warning(request, 'ابتدا شماره موبایل یا ایمیل خود را در پنل کاربری وارد کنید.')
            return redirect('core:check_view')
 
        form = LoginPasswordForm(request.POST)
        if form.is_valid():
            entered_password = form.cleaned_data['password']
            if check_password(entered_password, student.password):
                request.session.pop('pending_student_id', None)
                request.session['auth_student_id'] = student.pk
                return redirect('core:check_view')
            form.add_error('password', 'رمز عبور اشتباه است.')
 
        return render(request, self.template_name, {'form': form, 'mode': 'login', 'student': student})
 
class LogoutView(View):
    def get(self, request):
        request.session.pop('auth_student_id', None)
        request.session.pop('pending_student_id', None)
        messages.info(request, 'از حساب کاربری خارج شدید.')
        return redirect('core:check_view')
 
class SuccessView(TemplateView):
    template_name = 'success.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['name'] = self.request.session.get('student_name', '')
        context['age'] = self.request.session.get('student_age', '')
        context['phone'] = self.request.session.get('student_phone', '')
        context['reshte'] = self.request.session.get('student_reshte', '')
        context['school'] = self.request.session.get('student_school', '')
        context['city'] = self.request.session.get('student_city', '')
        context['moaref'] = self.request.session.get('student_moaref', '')
        return context

class StudentSessionRequiredMixin:
    def dispatch(self, request, *args, **kwargs):
        if not request.session.get('auth_student_id'):
            messages.warning(
                request,
                'برای دسترسی به این صفحه ابتدا وارد پنل کاربری خود شوید.'
            )
            return redirect('core:check_view')
        return super().dispatch(request, *args, **kwargs)
 
 
class ClassOfflineView(StudentSessionRequiredMixin, TemplateView):
    template_name = 'class_offline.html'
 
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'جلسه پیش‌نیاز'
        return context
 
 
class ClassOnlineView(StudentSessionRequiredMixin, TemplateView):
    template_name = 'class_online.html'
 
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        session_number = kwargs.get('session_number')
        DEFAULT_C_LINK = "https://www.aparat.com/metaAcademy/live"
        
        video_links = {v.session_id: {'video_url': v.video_url, 'chat_url': v.chat_url, 'is_live': v.is_live} for v in VideoLink.objects.all()}
        
        CLASS_SESSIONS = {
            1: {
                'title': 'جلسه ۱: شروع طوفانی',
                'date': '۱۹ مرداد',
                'desc': 'مقدمات برنامه‌نویسی، نصب پایتون، عملگرها، دریافت ورودی و مبانی پایتون',
                'video_url': '',
            },
            2: {
                'title': 'جلسه ۲: ساختمان‌های داده',
                'date': '۲۶ مرداد',
                'desc': 'لیست‌ها، دیکشنری‌ها، تاپل‌ها، متدهای پرکاربرد هرکدام',
                'video_url': '',
            },
            3: {
                'title': 'جلسه ۳: کنترل جریان برنامه',
                'date': '۲ شهریور',
                'desc': 'شرط و حلقه‌ها، پیمایش پرسرعت، حلقه‌های تودرتو، دستورات مربوط به شرط و حلقه، حل مسائل منطقی',
                'video_url': '',
            },
            4: {
                'title': 'جلسه ۴: توابع و شیءگرایی',
                'date': '۹ شهریور',
                'desc': 'تابع، ورودی و خروجی، ماژول‌ها، کلاس و آبجکت، شیءگرایی، ارث‌بری، چندریختی',
                'video_url': '',
            },
            5: {
                'title': 'جلسه ۵: پروژه‌های واقعی با پایتون',
                'date': '۱۶ شهریور',
                'desc': 'مدیریت و خواندن/نوشتن فایل، مدیریت خطاها، مفهوم استثناءها و مدیریت آنها، رمزنگاری، امنیت در پایتون، بازی‌سازی با پایگیم',
                'video_url': '',
            },
        }
        
        for session_id in CLASS_SESSIONS:
            if session_id in video_links:
                CLASS_SESSIONS[session_id]['video_url'] = video_links[session_id]['video_url']
                CLASS_SESSIONS[session_id]['chat_url'] = video_links[session_id]['chat_url'] or DEFAULT_C_LINK
                CLASS_SESSIONS[session_id]['is_live'] = video_links[session_id]['is_live']
                        
        session_data = CLASS_SESSIONS.get(session_number)
 
        if not session_data:
            raise Http404('جلسه‌ی مورد نظر پیدا نشد.')
 
        context['session_number'] = session_number
        context['session_data'] = session_data
        context['all_sessions'] = CLASS_SESSIONS
        context['title'] = session_data['title']
        return context

class CertificateView(View):
    template_name = 'certificate.html'
    
    def get(self, request):
        # بررسی احراز هویت کاربر
        student_id = request.session.get('auth_student_id')
        if not student_id:
            messages.error(request, 'لطفاً ابتدا وارد سیستم شوید')
            return redirect('core:check')
        
        try:
            student = Student.objects.get(pk=student_id)
        except Student.DoesNotExist:
            request.session.pop('auth_student_id', None)
            messages.error(request, 'کاربر یافت نشد')
            return redirect('core:check')
        
        certificate_url = None
        if student.certificate_file:
            certificate_url = student.certificate_file.url
            
        # پر کردن فرم با اطلاعات فعلی دانشجو
        initial_data = {
            'name': student.name,
            'national_code': student.national_code,
        }
        form = CertificateForm(initial=initial_data)
        
        return render(request, self.template_name, {
            'form': form,
            'student': student,
            'certificate_url': certificate_url,
        })
    
    def post(self, request):
        # بررسی احراز هویت
        student_id = request.session.get('auth_student_id')
        if not student_id:
            messages.error(request, 'لطفاً ابتدا وارد سیستم شوید')
            return redirect('core:check')
        
        try:
            student = Student.objects.get(pk=student_id)
        except Student.DoesNotExist:
            request.session.pop('auth_student_id', None)
            messages.error(request, 'کاربر یافت نشد')
            return redirect('core:check')
        
        if student.certificate_file and student.is_certified:
            messages.error(request, '❌ مدرک شما قبلاً ساخته شده است و قابل ویرایش نیست.')
            return redirect('core:certificate')
        
        form = CertificateForm(request.POST)
        
        if form.is_valid():
            # ذخیره اطلاعات در دیتابیس
            student.name = form.cleaned_data['name']
            student.national_code = form.cleaned_data['national_code']
            student.save()
            
            if student.is_certified:
                try:
                    cert_content = generate_certificate_for_student(student)
                    student.certificate_file.save(cert_content.name, cert_content, save=True)
                    messages.success(request, '✅ مدرک شما آماده شد.')
                except Exception as e:
                    messages.error(request, f'خطا در ساخت مدرک: هنوز واجد شرایط دریافت گواهی نیستید!.')
            else:
                return redirect('core:payment')

            return redirect('core:certificate')

        # در صورت نامعتبر بودن فرم
        return render(request, self.template_name, {
            'form': form,
            'student': student,
        })
        
def admin_dashboard(request):
    # ========== آمار ==========
    total_students = Student.objects.count()
    has_signature = Signature.objects.exists()
    
     # ========== پردازش آپلود امضا ==========
    if request.method == 'POST' and request.FILES.get('signature_image'):
        sig, created = Signature.objects.get_or_create(user=request.user)
        sig.image = request.FILES['signature_image']
        sig.save()
        messages.success(request, '✅ امضا با موفقیت آپلود شد')
        return redirect('core:admin_dashboard')
    
    # ========== ساخت پیشنمایش (فقط روی قالب خالی) ==========
    preview_image_url = None
    signature = Signature.objects.first()
    
    if signature:
        try:
            preview_image_url = preview_signature_on_template(
                signature.image.path,
                'static/img/certificate_template.jpg'
            )
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"خطای پیشنمایش: {error_detail}")
            messages.error(request, f'خطا در ساخت پیشنمایش: {e}')
    
    context = {
        'total_students': total_students,
        'has_signature': has_signature,
        'signature': signature,
        'preview_image_url': preview_image_url,
    }
    return render(request, 'admin_dashboard.html', context)

def payment_request_view(request):
    student_id = request.session.get('auth_student_id')
    if not student_id:
        messages.error(request, 'لطفاً ابتدا وارد سیستم شوید')
        return redirect('core:check')

    try:
        student = Student.objects.get(pk=student_id)
    except Student.DoesNotExist:
        request.session.pop('auth_student_id', None)
        messages.error(request, 'کاربر یافت نشد')
        return redirect('core:check')

    # اگر قبلاً تایید شده
    if student.is_certified:
        messages.info(request, 'شما قبلاً پرداخت خود را ثبت کرده‌اید و مدرک شما فعال است.')
        return redirect('core:certificate')

    # اگر قبلاً درخواست داده ولی هنوز تایید نشده
    if hasattr(student, 'payment_request'):
        messages.warning(request, 'درخواست شما قبلاً ثبت شده و در انتظار تأیید است.')
        return redirect('core:certificate')

    if request.method == 'POST':
        form = PaymentForm(request.POST)
        
        if form.is_valid():
            payment = form.save(commit=False)
            payment.student = student
            payment.save()
            messages.success(request, '✅ درخواست شما ثبت شد. پس از تأیید واحد حسابداری، مدرک شما فعال می‌شود.')
            return redirect('core:certificate')
    else:
        form = PaymentForm()

    return render(request, 'payment.html', {
        'form': form,
        'student': student
    })
    
@staff_member_required
def export_excel(request):
    """
    خروجی اکسل از تمام دانش‌آموزان (همه فیلدها به صورت متن)
    """
    students = Student.objects.all().order_by('-created_at')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'دانش‌آموزان'
    
    # استایل‌ها
    header_font = Font(name='Bidad', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    cell_font = Font(name='Bidad', size=11)
    cell_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # هدرها
    headers = ['ردیف', 'نام و نام خانوادگی', 'سن', 'شماره تلفن', 'کدملی', 'رشته تحصیلی', 'مدرسه', 'شهر', 'معرف', 'تاریخ ثبت']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # داده‌ها (همه به صورت مستقیم)
    for row, student in enumerate(students, 2):
        ws.cell(row=row, column=1, value=row-1).border = border
        ws.cell(row=row, column=2, value=student.name).border = border
        ws.cell(row=row, column=3, value=student.age).border = border
        ws.cell(row=row, column=4, value=student.phone_number).border = border
        ws.cell(row=row, column=5, value=student.national_code).border = border
        ws.cell(row=row, column=6, value=student.reshte).border = border  # ← مستقیم و بدون تغییر
        ws.cell(row=row, column=7, value=student.school).border = border
        ws.cell(row=row, column=8, value=student.city).border = border
        ws.cell(row=row, column=9, value=student.moaref or '').border = border
        created_at_local = timezone.localtime(student.created_at)
        ws.cell(row=row, column=10, value=created_at_local.strftime('%Y/%m/%d %H:%M')).border = border
        
        for col in range(1, 10):
            ws.cell(row=row, column=col).font = cell_font
            ws.cell(row=row, column=col).alignment = cell_alignment
    
    # عرض ستون‌ها
    column_widths = {
        'A': 8, 'B': 25, 'C': 10, 'D': 18, 
        'E': 18, 'F': 25, 'G': 25, 'H': 15, 'I': 20, 'J': 20
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    ws.row_dimensions[1].height = 30
    for row in range(2, len(students) + 2):
        ws.row_dimensions[row].height = 25
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=students_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    
    wb.save(response)
    return response

@staff_member_required
def import_excel(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, '❌ فرمت فایل باید .xlsx یا .xls باشد!')
                return redirect('core:import_excel')
            
            try:
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                
                # خواندن هدرها (ردیف اول)
                headers = [cell.value for cell in ws[1]]
                
                # پیدا کردن اندیس ستون‌ها
                col_index = {}
                for idx, header in enumerate(headers):
                    if header:
                        header_str = str(header).strip()
                        if 'نام' in header_str:
                            col_index['name'] = idx
                        elif 'سن' in header_str:
                            col_index['age'] = idx
                        elif 'تلفن' in header_str or 'شماره' in header_str:
                            col_index['phone'] = idx
                        elif 'کدملی' in header_str:
                            col_index['national_code'] = idx
                        elif 'رشته' in header_str:
                            col_index['reshte'] = idx
                        elif 'مدرسه' in header_str:
                            col_index['school'] = idx
                        elif 'شهر' in header_str:
                            col_index['city'] = idx
                        elif 'معرف' in header_str:
                            col_index['moaref'] = idx
                        elif 'تاریخ' in header_str or 'ثبت' in header_str:
                            col_index['created_at'] = idx
                
                # بررسی وجود ستون‌های ضروری
                required = ['name', 'age', 'phone', 'reshte', 'school', 'city']
                for field in required:
                    if field not in col_index:
                        messages.error(request, f'❌ ستون "{field}" در فایل پیدا نشد!')
                        return redirect('core:import_excel')
                
                added_count = 0
                error_rows = []
                
                # خواندن داده‌ها از ردیف دوم به بعد
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or not any(row):
                        continue
                    
                    try:
                        name = str(row[col_index['name']]).strip() if row[col_index['name']] else ''
                        age = int(row[col_index['age']]) if row[col_index['age']] else None
                        phone_number = str(row[col_index['phone']]).strip() if row[col_index['phone']] else ''
                        national_code = str(row[col_index['national_code']]).strip() if row[col_index['national_code']] else ''
                        reshte = str(row[col_index['reshte']]).strip() if row[col_index['reshte']] else ''
                        school = str(row[col_index['school']]).strip() if row[col_index['school']] else ''
                        city = str(row[col_index['city']]).strip() if row[col_index['city']] else ''
                        moaref = str(row[col_index.get('moaref')]).strip() if col_index.get('moaref') and row[col_index['moaref']] else None
                        created_at_str = str(row[col_index.get('created_at')]).strip() if col_index.get('created_at') and row[col_index['created_at']] else None
                        
                        # اعتبارسنجی
                        if not name:
                            error_rows.append(f'ردیف {row_idx}: نام نمی‌تواند خالی باشد')
                            continue
                        if age is None:
                            age = 1
                        else:
                            try:
                                age = int(age)
                                if age < 1 or age > 120:
                                    error_rows.append(f'ردیف {row_idx}: سن باید بین 1 تا 120 باشد')
                                    continue
                            except (ValueError, TypeError):
                                error_rows.append(f'ردیف {row_idx}: سن باید عدد باشد')
                                continue
                        # شماره تلفن - تبدیل به رشته و استانداردسازی
                        phone_number = str(phone_number).strip() if phone_number else ''
                        # اگر با 0 شروع نمی‌شه و 10 رقمه، 0 رو اولش بذار
                        if phone_number and not phone_number.startswith('0') and len(phone_number) == 10:
                            phone_number = '0' + phone_number
                        if not phone_number or not phone_number.startswith('09') or len(phone_number) != 11:
                            error_rows.append(f'ردیف {row_idx}: شماره تلفن باید با 09 شروع شود و 11 رقم باشد')
                            continue
                        if not reshte:
                            reshte = 'ثبت نشده'
                        if not school:
                            school = 'ثبت نشده'
                        if not city:
                            city = 'ثبت نشده'
                        
                        # ایجاد شیء دانش‌آموز
                        student = Student(
                            name=name,
                            age=age,
                            phone_number=phone_number,
                            national_code=national_code,
                            reshte=reshte,
                            school=school,
                            city=city,
                            moaref=moaref if moaref else None
                        )
                        try:
                            student.full_clean()
                        except ValidationError as e:
                            error_rows.append(f'ردیف {row_idx}: خطای اعتبارسنجی - {", ".join(e.messages)}')
                            continue

                        student.save()
                        
                        # اگر تاریخ ثبت در فایل وجود دارد، آن را تنظیم کن
                        if created_at_str:
                            try:
                                # تبدیل تاریخ از فرمت اکسل به datetime
                                # فرمت: 2026/07/04 23:07
                                created_at_dt = datetime.strptime(created_at_str, '%Y/%m/%d %H:%M')
                                
                                # اگر timezone فعال است، آن را aware کنید
                                if timezone.is_naive(created_at_dt):
                                    created_at_dt = timezone.make_aware(created_at_dt)
                                
                                # به‌روزرسانی فیلد created_at
                                Student.objects.filter(pk=student.pk).update(created_at=created_at_dt)
                                
                            except ValueError as e:
                                error_rows.append(f'ردیف {row_idx}: فرمت تاریخ صحیح نیست (مثال: 2026/07/04 23:07) - {str(e)}')
                        
                        added_count += 1
                        
                    except IntegrityError:
                        error_rows.append(f'ردیف {row_idx}: شماره تلفن {phone_number} تکراری است')
                    except Exception as e:
                        error_rows.append(f'ردیف {row_idx}: خطا - {str(e)}')
                
                # نمایش نتیجه
                if added_count > 0:
                    messages.success(request, f'✅ {added_count} دانش‌آموز با موفقیت اضافه شدند!')
                if error_rows:
                    for error in error_rows[:5]:
                        messages.warning(request, f'⚠️ {error}')
                    if len(error_rows) > 5:
                        messages.info(request, f'و {len(error_rows) - 5} خطای دیگر وجود دارد.')
                
                return redirect('core:students')
                
            except Exception as e:
                messages.error(request, f'❌ خطا در خواندن فایل: {str(e)}')
                return redirect('core:import_excel')
        else:
            messages.error(request, '❌ فرمت فایل صحیح نیست!')
    else:
        form = ExcelUploadForm()
    
    return render(request, 'import_excel.html', {'form': form})

